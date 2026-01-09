import streamlit as st
import pandas as pd
import io
from openpyxl.worksheet.datavalidation import DataValidation
import database
import sqlite3
import random
from datetime import datetime, timedelta

# Regras globais de qualidade
REGRAS_QUALIDADE = {
    "min_experientes_por_turno": 1,
    "niveis_experientes": ["Senior", "Especialista", "Pleno"]
}


@st.cache_data(ttl=60)
def carregar_dados_locais():
    database.init_all_db_tables()
    conn = database.get_db_connection()
    try:
        # Seleciona colunas incluindo as preferências
        query = """
                SELECT id, \
                       nome, \
                       email, \
                       nivel, \
                       data_admissao, \
                       ativo, \
                       skill_cplug, \
                       skill_dd, \
                       pref_dia, \
                       pref_turno
                FROM analistas
                WHERE ativo = 1
                ORDER BY nome \
                """
        df_analistas = pd.read_sql_query(query, conn)
        # Remove colunas duplicadas se houver
        df_analistas = df_analistas.loc[:, ~df_analistas.columns.duplicated()]

        df_indisp = pd.read_sql_query("SELECT * FROM indisponibilidades", conn)
        conn.close()
        return df_analistas, df_indisp
    except Exception as e:
        # st.error(f"Erro ao carregar dados: {e}") # Debug se necessário
        try:
            conn.close()
        except:
            pass
        return pd.DataFrame(), pd.DataFrame()


def load_staff_rules_from_db():
    conn = database.get_db_connection()
    regras = {}

    # --- MUDANÇA AQUI: Feriado agora segue o padrão de Sábado (5, 4, 1) ---
    padrao = {
        "Sabado": {"Manha": 5, "Noite": 4, "Integral": 1},
        "Domingo": {"Manha": 4, "Noite": 3, "Integral": 1},
        "Feriado": {"Manha": 5, "Noite": 4, "Integral": 1}  # Ajustado
    }
    # ----------------------------------------------------------------------

    try:
        df = pd.read_sql_query("SELECT dia_tipo, turno, quantidade FROM regras_staff", conn)
        if df.empty: return padrao

        for _, row in df.iterrows():
            if row['dia_tipo'] not in regras: regras[row['dia_tipo']] = {}
            regras[row['dia_tipo']][row['turno']] = row['quantidade']

        # Garante que chaves faltantes usem o padrão
        for dia, turnos in padrao.items():
            if dia not in regras:
                regras[dia] = turnos
            else:
                for turno, qtd in turnos.items():
                    if turno not in regras[dia]: regras[dia][turno] = qtd

        return regras
    except:
        return padrao
    finally:
        conn.close()


def load_shift_hours_from_db():
    conn = database.get_db_connection()
    padrao = {"Manha": 5.5, "Noite": 5.0, "Integral": 10.0}
    try:
        df = pd.read_sql_query("SELECT turno, horas FROM configuracao_turnos", conn)
        if df.empty: return padrao
        return dict(zip(df['turno'], df['horas']))
    except:
        return padrao
    finally:
        conn.close()


def load_max_hours_limit():
    conn = database.get_db_connection()
    try:
        df = pd.read_sql_query("SELECT valor FROM configuracao_limites WHERE chave='max_horas_ciclo'", conn)
        if df.empty: return 30.0
        return float(df.iloc[0]['valor'])
    except:
        return 30.0
    finally:
        conn.close()


def to_excel(df):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Escala', index=True)
        ws = writer.sheets['Escala']

        # Adiciona validação de dados (Dropdown) nas células do Excel
        opcoes = ["FOLGA", "Manha", "Noite", "Integral", "Ferias"]
        formula = f'"{",".join(opcoes)}"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)

        # Aplica validação da célula B2 até o fim da tabela
        if ws.max_row > 2:  # Garante que tem dados além do cabeçalho e mentor/sobreaviso
            # Ajuste dinâmico para pegar a área correta dos analistas
            # Considerando que as últimas 2 linhas podem ser Mentor/Sobreaviso,
            # a validação vai até max_row - 2 (ajuste conforme necessidade)
            dv.add(f"B2:{ws.cell(ws.max_row - 2, ws.max_column).coordinate}")
            ws.add_data_validation(dv)

    return output.getvalue()